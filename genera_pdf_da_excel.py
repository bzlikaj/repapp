"""
Script per generare PDF da un file Excel modificato
Legge le assegnazioni dall'Excel e crea un PDF
"""

import sys
import os
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# Aggiungi il percorso src al path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

from pdf_generator import PDFCalendarioGenerator


class CalendarioFromExcel:
    """Classe per leggere assegnazioni da Excel e generare PDF"""
    
    def __init__(self, excel_path):
        """
        Inizializza dal file Excel
        
        Args:
            excel_path: percorso del file Excel modificato
        """
        self.excel_path = excel_path
        self.anno = 2026
        self.assegnazioni = {}  # {"2026-01-01": ("Tecnico", "tipo"), ...}
        self.TECNICI = [
            "Likaj", "Ferraris", "Zanotto", "Casazza", "Mancin",
            "Dardha", "Franchini", "Giraldin", "Terazzi"
        ]
        self.FESTIVI_2026 = [
            "2026-01-01", "2026-01-06", "2026-04-12", "2026-04-13",
            "2026-04-25", "2026-05-01", "2026-06-02", "2026-08-15",
            "2026-11-01", "2026-12-08", "2026-12-25", "2026-12-26"
        ]
        self._leggi_excel()
    
    def _leggi_excel(self):
        """Legge le assegnazioni dal file Excel"""
        wb = load_workbook(self.excel_path)
        
        # Nomi dei mesi negli sheet
        nomi_mesi = [
            "gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
            "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"
        ]
        
        mesi_map = {nome.upper(): i+1 for i, nome in enumerate(nomi_mesi)}
        
        # Leggi ogni foglio mese
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() == "ISTRUZIONI":
                continue
            
            mese = mesi_map.get(sheet_name.upper())
            if not mese:
                continue
            
            ws = wb[sheet_name]
            
            # Leggi i dati dalla griglia
            # Header è a riga 3 (LUN, MAR, MER, GIO, VEN, SAB, DOM)
            riga_inizio = 4
            
            for riga_idx, riga in enumerate(ws.iter_rows(min_row=riga_inizio, max_row=50, values_only=False), start=riga_inizio):
                for col_idx, cell in enumerate(riga, start=1):
                    if cell.value is None:
                        continue
                    
                    # Estrai il numero del giorno dalla cella (primo numero)
                    cell_value = str(cell.value).strip()
                    if not cell_value:
                        continue
                    
                    # Dividi per newline se presente
                    linee = cell_value.split('\n')
                    if len(linee) < 2:
                        continue
                    
                    try:
                        giorno = int(linee[0].strip())
                        tecnico = linee[1].strip() if len(linee) > 1 else ""
                    except (ValueError, IndexError):
                        continue
                    
                    if giorno < 1 or giorno > 31:
                        continue
                    
                    if not tecnico or tecnico not in self.TECNICI:
                        continue
                    
                    # Determina il tipo
                    data_str = f"{self.anno}-{mese:02d}-{giorno:02d}"
                    if data_str in self.FESTIVI_2026:
                        tipo = "festivo"
                    elif col_idx in [6, 7]:  # Sabato o domenica
                        tipo = "weekend"
                    else:
                        tipo = "feriale"
                    
                    self.assegnazioni[data_str] = (tecnico, tipo)
        
        print(f"✅ Letti {len(self.assegnazioni)} giorni dal file Excel")
    
    def get_reperibile_data(self, data_str):
        """Ottiene il reperibile per una data"""
        return self.assegnazioni.get(data_str, ("", ""))


def main():
    """Funzione principale"""
    print("=" * 60)
    print("GENERATORE PDF DA EXCEL MODIFICATO")
    print("=" * 60)
    
    # Percorso dell'Excel
    excel_path = Path(__file__).parent / "output" / "calendario_reperibilita_2026.xlsx"
    
    if not excel_path.exists():
        print(f"❌ Errore: File Excel non trovato in {excel_path}")
        return
    
    print(f"\n1️⃣ Lettura Excel...")
    print(f"   Percorso: {excel_path}")
    
    try:
        calendario = CalendarioFromExcel(str(excel_path))
    except Exception as e:
        print(f"❌ Errore nella lettura dell'Excel: {e}")
        return
    
    # Genera il PDF
    print(f"\n2️⃣ Generazione PDF...")
    output_dir = Path(__file__).parent / "output"
    
    # Crea nome file con data/ora per non sovrascrivere
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_path = str(output_dir / f"calendario_reperibilita_2026_{timestamp}.pdf")
    
    try:
        pdf_generator = PDFCalendarioGenerator(calendario, pdf_path)
        pdf_generator.genera_pdf()
        print("✅ PDF generato con successo!")
    except Exception as e:
        print(f"❌ Errore nella generazione del PDF: {e}")
        return
    
    print("\n" + "=" * 60)
    print(f"📄 File PDF salvato in:")
    print(f"   {pdf_path}")
    print("=" * 60)


if __name__ == "__main__":
    main()

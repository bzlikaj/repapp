"""
Generatore di Excel per il calendario di reperibilità
Consente modifica manuale dei nomi dei tecnici
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import calendar


class GeneratoreExcel:
    """Genera file Excel del calendario di reperibilità"""
    
    # Colori
    COLORI = {
        "festivo": "FF4757",      # Rosso vivace
        "weekend": "00B4DB",       # Ciano vivace
        "feriale": "F8F9FA",       # Grigio chiaro
        "header": "2C3E50",        # Grigio scuro
    }
    
    COLORI_TESTO = {
        "festivo": "FFFFFF",       # Bianco
        "weekend": "FFFFFF",       # Bianco
        "feriale": "333333",       # Grigio scuro
        "header": "FFFFFF",        # Bianco
    }
    
    def __init__(self, calendario):
        """
        Inizializza il generatore
        
        Args:
            calendario: istanza di CalendarioReperibilita
        """
        self.calendario = calendario
        self.anno = int(getattr(calendario, "anno", 2026) or 2026)
        # Cache assegnazioni per allineare Excel a UI/API (tecnico, tipo, aiutante)
        try:
            self._assegnazioni = dict(getattr(calendario, "assegnazioni", {}) or {})
        except Exception:
            self._assegnazioni = {}
        if hasattr(calendario, "get_festivi"):
            try:
                self.festivi = set(calendario.get_festivi(self.anno))
            except Exception:
                self.festivi = set()
        else:
            self.festivi = set(getattr(calendario, "FESTIVI_2026", []) or [])
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
    
    def genera_excel(self, filepath):
        """
        Genera il file Excel
        
        Args:
            filepath: percorso del file di output
        """
        # Crea foglio di istruzioni
        self._crea_foglio_istruzioni()
        
        # Crea un foglio per ogni mese
        for mese in range(1, 13):
            self._crea_foglio_mese(mese)
        
        # Salva (filepath può essere un path o un file-like object)
        self.wb.save(filepath)
        if isinstance(filepath, (str, bytes)):
            print(f"Excel generato: {filepath}")
    
    def _crea_foglio_istruzioni(self):
        """Crea il foglio con le istruzioni"""
        ws = self.wb.create_sheet("ISTRUZIONI", 0)
        
        # Titolo
        ws['A1'] = f"CALENDARIO DI REPERIBILITÀ {self.anno}".upper()
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        riga = 3
        
        # SEZIONE 1: ISTRUZIONI BASE
        ws[f'A{riga}'] = "ISTRUZIONI D'USO:".upper()
        ws[f'A{riga}'].font = Font(size=12, bold=True)
        riga += 1
        
        istruzioni_base = [
            f"• Ogni foglio rappresenta un mese del {self.anno}",
            "• Modifica i nomi dei tecnici direttamente nelle celle colorate",
            "• Usa CTRL+S per salvare il file dopo le modifiche",
            "• I giorni sono organizzati per settimana (lunedì-domenica)",
            "• Domenica è sempre assegnata allo stesso tecnico del sabato",
            "• Il numero in alto-sinistra è il giorno del mese",
        ]
        
        for testo in istruzioni_base:
            ws[f'A{riga}'] = (testo or "").upper()
            ws[f'A{riga}'].font = Font(size=10)
            ws[f'A{riga}'].alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[riga].height = 20
            riga += 1
        
        riga += 1
        
        # SEZIONE 2: TIPI DI REPERIBILITÀ
        ws[f'A{riga}'] = "TIPI DI REPERIBILITÀ:".upper()
        ws[f'A{riga}'].font = Font(size=12, bold=True)
        riga += 1
        
        tipologie = [
            ("FERIALE", "Lunedì-venerdì: singolo giorno di reperibilità"),
            ("WEEKEND", "Sabato + domenica: weekend completo assegnato insieme"),
            ("FESTIVO", f"Giorni festivi nazionali ({len(getattr(self.calendario, 'get_festivi', lambda a: [])(self.anno))} nel {self.anno})"),
        ]
        
        for tipo, desc in tipologie:
            ws[f'A{riga}'] = f"• {tipo}".upper()
            ws[f'A{riga}'].font = Font(size=10, bold=True)
            riga += 1
            ws[f'A{riga}'] = f"  {desc}".upper()
            ws[f'A{riga}'].font = Font(size=10)
            ws[f'A{riga}'].alignment = Alignment(wrap_text=True)
            riga += 1
        
        riga += 1
        
        # SEZIONE 3: REGOLE IMPORTANTI
        ws[f'A{riga}'] = "REGOLE IMPORTANTI:".upper()
        ws[f'A{riga}'].font = Font(size=12, bold=True)
        riga += 1
        
        regole = [
            "ROTAZIONE EQUA: I 9 tecnici si alternano in modo giusto e bilanciato",
            "REGOLA 7 GIORNI: Dopo un weekend o festivo, il tecnico non può avere altri turni",
            "  per 7 giorni prima e 7 giorni dopo",
            "TUTTI LAVORANO: Ogni tecnico ha ~40 turni l'anno (38-43 a causa dei vincoli)",
            ("1° GENNAIO: Nel 2026 è assegnato a Dardha (obbligo aziendale)" if self.anno == 2026
             else "1° GENNAIO: Negli altri anni segue la rotazione normale (nessun obbligo)"),
        ]
        
        for testo in regole:
            ws[f'A{riga}'] = (testo or "").upper()
            ws[f'A{riga}'].font = Font(size=10)
            ws[f'A{riga}'].alignment = Alignment(wrap_text=True)
            ws.row_dimensions[riga].height = 25
            riga += 1
        
        riga += 1
        
        # SEZIONE 4: LEGENDA COLORI
        ws[f'A{riga}'] = "LEGENDA COLORI:".upper()
        ws[f'A{riga}'].font = Font(size=12, bold=True)
        riga += 1
        
        legenda = [
            ("Festivo", "FF4757", "FFFFFF", "Giorni festivi nazionali (Natale, Pasqua, ecc.)"),
            ("Weekend", "00B4DB", "FFFFFF", "Sabato e domenica assegnati insieme"),
            ("Feriale", "F8F9FA", "333333", "Giorni lavorativi da lunedì a venerdì"),
        ]
        
        for tipo, bg_color, text_color, desc in legenda:
            cell_tipo = f'A{riga}'
            ws[cell_tipo] = (tipo or "").upper()
            ws[cell_tipo].fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            ws[cell_tipo].font = Font(color=text_color, bold=True, size=11)
            ws[cell_tipo].alignment = Alignment(horizontal="center")
            ws.column_dimensions['A'].width = 15
            
            cell_desc = f'B{riga}'
            ws[cell_desc] = (desc or "").upper()
            ws[cell_desc].font = Font(size=10)
            ws[cell_desc].alignment = Alignment(wrap_text=True)
            ws.column_dimensions['B'].width = 65
            ws.row_dimensions[riga].height = 25
            riga += 1
        
        riga += 1
        
        # SEZIONE 5: COME USARE
        ws[f'A{riga}'] = "COME USARE QUESTO FILE:".upper()
        ws[f'A{riga}'].font = Font(size=12, bold=True)
        riga += 1
        
        procedura = [
            "1. Apri questo file Excel",
            "2. Vai al foglio del mese che vuoi modificare",
            "3. Fai DOPPIO CLIC sulla cella con il nome del tecnico",
            "4. Modifica il nome e premi INVIO per confermare",
            "5. Ripeti per tutti i giorni che vuoi cambiare",
            "6. Premi CTRL+S per salvare le modifiche",
            "7. Stampa o condividi il file aggiornato",
        ]
        
        for testo in procedura:
            ws[f'A{riga}'] = (testo or "").upper()
            ws[f'A{riga}'].font = Font(size=10)
            ws[f'A{riga}'].alignment = Alignment(wrap_text=True)
            riga += 1
        
        riga += 1
        
        # SEZIONE 6: TECNICI
        ws[f'A{riga}'] = "I 9 TECNICI:".upper()
        ws[f'A{riga}'].font = Font(size=12, bold=True)
        riga += 1
        
        tecnici = self.calendario.TECNICI
        for i, tecnico in enumerate(tecnici, 1):
            ws[f'A{riga}'] = f"{i}. {(tecnico or '').upper()}"
            ws[f'A{riga}'].font = Font(size=10)
            if i % 3 == 0:
                riga += 1
            else:
                ws[f'A{riga}'].alignment = Alignment(horizontal="left")
                ws.merge_cells(f'A{riga}:B{riga}')
        
        riga += 2
        
        # SEZIONE 7: CONTATTI
        ws[f'A{riga}'] = "NOTE FINALI:"
        ws[f'A{riga}'].font = Font(size=11, bold=True, italic=True)
        riga += 1
        
        note = [
            "• Se trovi errori o incongruenze, segnala subito",
            "• La differenza tra i turni di ogni tecnico è dovuta ai vincoli rigidi",
            "• Conserva una copia del file originale prima di modificare",
            "• Allega il file aggiornato alle comunicazioni ufficiali",
        ]
        
        for testo in note:
            ws[f'A{riga}'] = (testo or "").upper()
            ws[f'A{riga}'].font = Font(size=9, italic=True)
            ws[f'A{riga}'].alignment = Alignment(wrap_text=True)
            riga += 1
    
    def _crea_foglio_mese(self, mese):
        """
        Crea il foglio per un mese
        
        Args:
            mese: numero del mese (1-12)
        """
        nome_mese = [
            "gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
            "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"
        ][mese - 1]
        
        ws = self.wb.create_sheet(nome_mese.upper())
        
        # Titolo
        titolo = f"REPERIBILITÀ {nome_mese.upper()} {self.anno}"
        ws['A1'] = titolo
        ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25
        
        # Header settimane
        giorni_settimana = ["LUN", "MAR", "MER", "GIO", "VEN", "SAB", "DOM"]
        for col, giorno in enumerate(giorni_settimana, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = giorno
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[3].height = 20
        
        # Larghezze colonne
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 22
        
        # Ottieni tutti i giorni del mese
        primo_giorno = datetime(self.anno, mese, 1)
        ultimo_giorno = calendar.monthrange(self.anno, mese)[1]
        
        # Determina il giorno della settimana del primo giorno (0=lunedì, 6=domenica)
        primo_weekday = primo_giorno.weekday()
        
        # Riempi la tabella
        riga_attuale = 4
        giorno_attuale = 1
        
        # Aggiungi righe vuote iniziali se il mese non inizia di lunedì
        if primo_weekday > 0:
            riga_attuale += 1
        
        while giorno_attuale <= ultimo_giorno:
            data = datetime(self.anno, mese, giorno_attuale)
            weekday = data.weekday()
            
            # Calcola la colonna (0=lunedì, 6=domenica)
            col = weekday + 1
            
            # Aggiungi una nuova riga se necessario
            if weekday == 0 and giorno_attuale > 1:
                riga_attuale += 1
            
            # Ottieni le informazioni di assegnazione
            data_str = data.strftime("%Y-%m-%d")
            tecnico = ""
            tipo = ""
            aiutante = ""

            assegnazione = self._assegnazioni.get(data_str)
            if isinstance(assegnazione, (list, tuple)) and len(assegnazione) >= 2:
                tecnico = assegnazione[0] or ""
                tipo = assegnazione[1] or ""
                if len(assegnazione) >= 3:
                    aiutante = assegnazione[2] or ""
            else:
                tecnico, tipo = self.calendario.get_reperibile_data(data_str)
                if hasattr(self.calendario, "get_aiutante_data"):
                    try:
                        aiutante = self.calendario.get_aiutante_data(data_str) or ""
                    except Exception:
                        aiutante = ""
            
            # Determina il colore (se la festività cade nel weekend, rimane "weekend")
            if data_str in self.festivi and tipo != "weekend":
                tipo_colore = "festivo"
            elif tipo == "weekend":
                tipo_colore = "weekend"
            else:
                tipo_colore = "feriale"
            
            # Crea la cella
            cell = ws.cell(row=riga_attuale, column=col)
            tecnico_up = (tecnico or "").strip().upper()
            aiutante_up = (aiutante or "").strip().upper()
            cell_value = f"{giorno_attuale}"
            if tecnico_up:
                cell_value += f"\n{tecnico_up}"
            if aiutante_up:
                cell_value += f"\n{aiutante_up}"
            cell.value = cell_value
            
            # Applica stile
            bg_color = self.COLORI[tipo_colore]
            text_color = self.COLORI_TESTO[tipo_colore]
            
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.font = Font(color=text_color, size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Border
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = border
            
            # Altezza riga (giorno + tecnico + aiutante)
            ws.row_dimensions[riga_attuale].height = 65
            
            giorno_attuale += 1
        
        # Applica border a tutte le celle vuote della griglia
        for riga in range(4, riga_attuale + 1):
            for col in range(1, 8):
                cell = ws.cell(row=riga, column=col)
                if cell.value is None:
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = border
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
